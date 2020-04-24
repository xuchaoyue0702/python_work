#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @Author: ccyy
# @Date: 2020/4/23

import openpyxl

# 行：row
# 列：column
# 单元格：cell
# 表：sheet

# 只能打开存在的表格，不能用该方法创新一个新的表格
workbook = openpyxl.load_workbook('C:/Users/17717/Desktop/OPPO有效故障_2019.xlsx')
print(workbook.sheetnames)
print(workbook['Sheet1'].dimensions)
sheet = workbook.active
cell = sheet['B4']
print(cell.value)
print(cell.row, cell.column, cell.coordinate)
cell2 = sheet.cell(row=2, column=3)
# 获取一列好几个
cells = sheet['A1:A3']
# 获取1列
cells2 = sheet['A']
# 获取好几列
cells3 = sheet['A:C']
# 获取第几行
cells4 = sheet[5]
for row in sheet.rows:
    # for cell in row:
    #     print(cell.value)
    print(row)

sheet['A1'] = 'hahah'
cell.value = 'xuxuxu'
workbook.save('C:/Users/17717/Desktop/OPPO有效故障_2019.xlsx')

